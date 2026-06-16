import openpyxl

template_path = r'C:\Users\DenisG\Documents\Workplace\Data\Prices\Line up monitoring.xlsx'
wb = openpyxl.load_workbook(template_path)
sheet = wb.active

for row_idx in range(2, 5):
    price_type = sheet.cell(row=row_idx, column=2).value
    print(f"Row {row_idx}: price_type='{price_type}'")
    if not price_type or str(price_type).strip() not in ['Lowest Price', 'Lowest Price LFR']:
        print("Skipped")
        continue
    print("Writing to cell")
    sheet.cell(row=row_idx, column=3).value = 'N/A'
    
wb.save(r'C:\Users\DenisG\Documents\Workplace\Data\Prices\test_debug.xlsx')
