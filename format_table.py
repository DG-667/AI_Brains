import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

file_path = r'C:\Users\DenisG\Documents\Workplace\Data\GFK\GFK data TTL.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Data']

max_col = ws.max_column
max_row = ws.max_row
col_letter = get_column_letter(max_col)
ref = f"A1:{col_letter}{max_row}"

tab = Table(displayName="GFKDataTable", ref=ref)
style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
tab.tableStyleInfo = style

ws.add_table(tab)
wb.save(file_path)
print('Done!')
