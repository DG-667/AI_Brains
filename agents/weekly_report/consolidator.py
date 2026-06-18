import pandas as pd
import json
import os
import sys
import shutil
from datetime import datetime
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils import get_column_letter

# Папки
INPUT_DIR = r'C:\Users\DenisG\Documents\Workplace\Data'
OUTPUT_DIR = r'C:\Users\DenisG\Documents\Workplace\Report'
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Загрузка маппинга партнеров
with open('partners_map.json', 'r', encoding='utf-8') as f:
    PARTNERS_MAP = json.load(f)

# Целевые сегменты для фильтрации (Daily Ship, Dasi)
ALLOWED_SEGMENTS = ["consumer notebook", "consumer all-in-one", "consumer desktop"]

# Расширенный список для Wesi (включает Commercial)
WESI_SEGMENTS = ["consumer notebook", "consumer all-in-one", "consumer desktop",
                 "commercial notebook", "commercial all-in-one", "commercial desktop"]

def normalize_partner(name):
    if pd.isna(name):
        return "Unknown"
    name_str = str(name).strip().lower()
    for canonical, synonyms in PARTNERS_MAP.items():
        if canonical.lower() == name_str:
            return canonical
        for syn in synonyms:
            if syn.lower() in name_str: 
                return canonical
    return str(name).strip()

def find_file(keyword, extension=None):
    for root, dirs, files in os.walk(INPUT_DIR):
        for file in files:
            if file.startswith('~$'): continue
            if keyword.lower() in file.lower():
                if extension and not file.lower().endswith(extension):
                    continue
                return os.path.join(root, file)
    return None

def find_header_row(filepath, key_column):
    for i in range(10):
        try:
            df = pd.read_excel(filepath, skiprows=i, nrows=2)
            if key_column in df.columns or key_column.lower() in [str(c).lower() for c in df.columns]:
                return i
        except:
            pass
    return 0

def filter_by_segments(df, df_name, target_col):
    """Фильтрация по целевым продуктовым группам"""
    if target_col not in df.columns:
        print(f"Внимание: В файле {df_name} не найдена колонка {target_col}. Фильтрация по сегменту пропущена.")
        return df

    mask = df[target_col].astype(str).str.lower().str.strip().isin(ALLOWED_SEGMENTS)
    filtered_df = df[mask].copy()
    print(f"{df_name}: Отфильтровано по сегментам ({target_col}). Осталось строк: {len(filtered_df)} из {len(df)}")
    return filtered_df

def get_subsegment(category, description, cpu, video):
    import re
    if str(category).strip().lower() != 'consumer notebook':
        return None
    
    desc_upper = str(description).upper()
    cpu_upper = str(cpu).upper()
    vid_upper = str(video).upper()

    desc_cpu_str = f"{cpu_upper} {desc_upper}"

    is_celeron = bool(re.search(r'(?i)CELERON', desc_cpu_str))
    is_pentium_athlon = bool(re.search(r'(?i)PENTIUM|ATHLON', desc_cpu_str))
    is_i3 = bool(re.search(r'(?i)\bI3\b|ULTRA\s*3|\bU3\b', desc_cpu_str))
    is_r3 = bool(re.search(r'(?i)\bR3\b|RYZEN\s*3|AI_R3', desc_cpu_str))
    is_i5 = bool(re.search(r'(?i)\bI5\b|ULTRA\s*5|\bU5\b', desc_cpu_str))
    is_r5 = bool(re.search(r'(?i)\bR5\b|RYZEN\s*5|AI_R5', desc_cpu_str))
    is_i7 = bool(re.search(r'(?i)\bI7\b|ULTRA\s*7|\bU7\b|ULT7', desc_cpu_str))
    is_r7 = bool(re.search(r'(?i)\bR7\b|RYZEN\s*7|AI_R7', desc_cpu_str))
    is_i9 = bool(re.search(r'(?i)\bI9\b|ULTRA\s*9|\bU9\b|ULT9', desc_cpu_str))
    is_r9 = bool(re.search(r'(?i)\bR9\b|RYZEN\s*9|AI_R9', desc_cpu_str))
    is_snapdragon = bool(re.search(r'(?i)SNAPDRAGON', cpu_upper)) or bool(re.search(r'(?i)SD_X', desc_upper))

    # Используем границы слов \b чтобы избежать ложных срабатываний внутри других слов
    is_yoga = bool(re.search(r'(?i)\b(YOGA|YG|2N1)\b', desc_upper))
    is_gaming = bool(re.search(r'(?i)\b(LOQ|LEGION|LGN|LGNP5|LGN5|LGN7|LGN9)\b', desc_upper))

    if is_snapdragon:
        return 'SnapDr Laptops'

    if is_gaming:
        vid_match = re.search(r'(3050|4050|5050|3060|4060|5060|3070|4070|5070|3080|4080|5080|4090|5090)', vid_upper)
        vid_num = vid_match.group(1) if vid_match else ''
        
        prefix = 'Gaming Others'
        if is_i5 or is_r5:
            prefix = 'Gaming i5R5'
        elif is_i7 or is_r7:
            prefix = 'Gaming i7R7'
        elif is_i9 or is_r9:
            prefix = 'Gaming i9R9'

        if not vid_num:
            return 'Gaming Others'

        if vid_num in ['3050', '4050', '5050', '3060', '4060', '5060']:
            return f"{prefix} {vid_num}"
        elif vid_num in ['3070', '4070', '5070']:
            if prefix in ['Gaming i7R7', 'Gaming i9R9']:
                return f"{prefix} {vid_num}+"
            else:
                return f"{prefix} {vid_num}"
        elif vid_num in ['3080', '4080', '5080', '4090', '5090']:
            if vid_num in ['4080', '4090']: vid_mapped = '4080+'
            elif vid_num == '3080': vid_mapped = '3080+'
            elif vid_num in ['5080', '5090']: vid_mapped = '5070+'
            else: vid_mapped = f"{vid_num}+"
            return f"{prefix} {vid_mapped}"
        else:
            return 'Gaming Others'

    if is_yoga:
        if is_i3 or is_r3: return 'Premium i3R3'
        if is_i5 or is_r5: return 'Premium i5R5'
        if is_i7 or is_r7: return 'Premium i7R7'
        if is_i9 or is_r9: return 'Premium i9R9'
        return 'Premium i5R5' # fallback for yoga

    # Everything else in Consumer Notebook is Basic (Ideapad)
    if is_celeron: return 'Basic Celeron'
    if is_pentium_athlon: return 'Basic Pentium Athlon'
    if is_i3: return 'Basic i3'
    if is_r3: return 'Basic R3'
    if is_i5: return 'Basic i5'
    if is_r5: return 'Basic R5'
    if is_i7: return 'Basic i7'
    if is_r7: return 'Basic R7'
    if is_i9: return 'Basic i9'
    if is_r9: return 'Basic R9'
    
    return 'Basic i3' # fallback for missing cpu info


def load_workplace():
    path = find_file('WorkPlace', '.xlsx')
    if not path:
        print("Файл WorkPlace.xlsx не найден!")
        return None, None, None, None
    xl = pd.ExcelFile(path)
    
    target_sheet = None
    target_skip = 0
    
    for sheet in xl.sheet_names:
        for i in range(10):
            try:
                df = pd.read_excel(path, sheet_name=sheet, skiprows=i, nrows=2)
                if 'Part number' in df.columns or 'part number' in [str(c).lower() for c in df.columns]:
                    target_sheet = sheet
                    target_skip = i
                    break
            except:
                pass
        if target_sheet:
            break
            
    if not target_sheet:
        print("В файле WorkPlace.xlsx не найдена страница с данными (колонка Part number)!")
        return None, None, None, None
        
    df = pd.read_excel(path, sheet_name=target_sheet, skiprows=target_skip)
    return df, path, target_sheet, target_skip

def main(week_num):
    print(f"--- Сборка отчета за Неделю {week_num} ---")
    
    wp_df, wp_path, wp_sheet_name, wp_skip = load_workplace()
    if wp_df is None: return
    
    # Извлечение исторических цен
    historical_prices = {}
    if 'Price' in wp_df.columns and 'BP Name' in wp_df.columns:
        hist_df = wp_df.dropna(subset=['Price']).drop_duplicates(subset=['BP Name', 'Part number'], keep='last')
        historical_prices = dict(zip(zip(hist_df['BP Name'], hist_df['Part number']), hist_df['Price']))
    
    # Извлечение исторических данных по Segment
    historical_segments = {}
    if 'Segment' in wp_df.columns and 'BP Name' in wp_df.columns:
        seg_df = wp_df.dropna(subset=['Segment']).drop_duplicates(subset=['BP Name', 'Part number'], keep='last')
        historical_segments = dict(zip(zip(seg_df['BP Name'], seg_df['Part number']), seg_df['Segment']))
        
    # Извлечение исторических данных по Platform
    historical_platforms = {}
    if 'Platform' in wp_df.columns and 'BP Name' in wp_df.columns:
        plat_df = wp_df.dropna(subset=['Platform']).drop_duplicates(subset=['BP Name', 'Part number'], keep='last')
        historical_platforms = dict(zip(zip(plat_df['BP Name'], plat_df['Part number']), plat_df['Platform']))
        
    # Извлечение исторических спецификаций
    spec_cols_to_pull = ['Category', 'Description', 'CPU', 'CPU type', 'RAM', 'Storage', 'Video', 'OS']
    avail_spec_cols = [c for c in spec_cols_to_pull if c in wp_df.columns]
    
    historical_specs = {}
    if avail_spec_cols and 'Part number' in wp_df.columns:
        hist_spec_df = wp_df[['Part number'] + avail_spec_cols].dropna(subset=['Part number']).drop_duplicates(subset=['Part number'], keep='last')
        historical_specs = hist_spec_df.set_index('Part number').to_dict(orient='index')
    
    print("\nЗагрузка Daily Shipment...")
    ds_path = find_file('Daily Ship', '.xlsb')
    if ds_path:
        ds_df = pd.read_excel(ds_path, sheet_name="DATA", engine='pyxlsb')
        
        # Фильтрация по Казахстану
        if 'COUNTRY' in ds_df.columns:
            kazakhstan_mask = ds_df['COUNTRY'].astype(str).str.lower().str.contains('kazakhstan', na=False)
            ds_df = ds_df[kazakhstan_mask]
            print(f"Daily Ship: Отфильтровано по Казахстану. Строк: {len(ds_df)}")
        
        # Фильтрация по сегментам
        ds_df = filter_by_segments(ds_df, "Daily Ship", "PH1_NAME")
        
        ds_df['BP Name'] = ds_df['CUSTNAME'].apply(normalize_partner)
        ds_df['Part number'] = ds_df['ORDPART']
        
        # Извлечение цен из Daily Ship (с привязкой к Партнеру)
        if 'UNIT_VALUE' in ds_df.columns:
            ds_prices = ds_df[['BP Name', 'Part number', 'UNIT_VALUE']].rename(columns={'UNIT_VALUE': 'Price'}).dropna()
            historical_prices.update(dict(zip(zip(ds_prices['BP Name'], ds_prices['Part number']), ds_prices['Price'])))
        
        # Фильтрация партнеров для Daily Ship
        ds_df = ds_df[ds_df['BP Name'].isin(['Marvel', 'STN', 'Sulpak', 'Technodom'])]
        print(f"Daily Ship: Оставлены партнеры Marvel, STN, Arena S(Sulpak), Technodom. Строк: {len(ds_df)}")
        
        # Агрегируем все строки (общий QTY и первый STATUS)
        ds_agg = ds_df.groupby(['BP Name', 'Part number']).agg({'QTY': 'sum', 'STATUS': 'first'}).reset_index()
        
        # Отдельно считаем SHIP QTY для сравнения с Transit
        ds_ship = ds_df[ds_df['STATUS'].astype(str).str.upper() == 'SHIP']
        ds_ship_agg = ds_ship.groupby(['BP Name', 'Part number'])['QTY'].sum().reset_index()
        ds_ship_agg = ds_ship_agg.rename(columns={'QTY': 'SHIP_QTY'})
        
        ds_agg = ds_agg.merge(ds_ship_agg, on=['BP Name', 'Part number'], how='left')
        ds_agg['SHIP_QTY'] = ds_agg['SHIP_QTY'].fillna(0)
        
        # Также берём первый НЕ-SHIP статус (если есть)
        ds_non_ship = ds_df[ds_df['STATUS'].astype(str).str.upper() != 'SHIP']
        ds_non_ship_status = ds_non_ship.groupby(['BP Name', 'Part number'])['STATUS'].first().reset_index()
        ds_non_ship_status = ds_non_ship_status.rename(columns={'STATUS': 'NON_SHIP_STATUS'})
        ds_agg = ds_agg.merge(ds_non_ship_status, on=['BP Name', 'Part number'], how='left')
    else:
        ds_agg = pd.DataFrame(columns=['BP Name', 'Part number', 'QTY', 'STATUS', 'SHIP_QTY', 'NON_SHIP_STATUS'])

    print("\nЗагрузка Dasi...")
    dasi_path = find_file('ee9dcc8b') or find_file('Dasi')
    if dasi_path:
        dasi_df = pd.read_excel(dasi_path)
        dasi_df = filter_by_segments(dasi_df, "Dasi", "PH Level 1")
        
        # Извлечение спецификаций Dasi (CPU SEGMENT -> CPU, Processor -> CPU type)
        dasi_spec_cols = {'Short Description': 'Description', 'CPU SEGMENT': 'CPU', 'Processor': 'CPU type', 'Memory': 'RAM', 'HDD': 'Storage', 'GPU Description': 'Video', 'OS': 'OS', 'Common Series': 'Platform', 'PH Level 1': 'Category'}
        avail_dasi_specs = ['PartNo'] + [col for col in dasi_spec_cols.keys() if col in dasi_df.columns]
        dasi_specs = dasi_df[avail_dasi_specs].drop_duplicates('PartNo').rename(columns=dasi_spec_cols)
        dasi_specs = dasi_specs.rename(columns={'PartNo': 'Part number'})
        
        dasi_df['BP Name'] = dasi_df['BP Name'].apply(normalize_partner)
        dasi_df['Part number'] = dasi_df['PartNo']
        
        # Транзит из Dasi (только Consumer, только целевые партнёры)
        if 'TRANSIT' in dasi_df.columns:
            transit_agg = dasi_df.groupby(['BP Name', 'Part number'])['TRANSIT'].sum().reset_index()
            transit_agg = transit_agg.rename(columns={'TRANSIT': 'Volume'})
            transit_agg = transit_agg[transit_agg['BP Name'].isin(['Marvel', 'STN', 'Sulpak', 'Technodom'])]
        else:
            transit_agg = pd.DataFrame(columns=['BP Name', 'Part number', 'Volume'])
        
        # СТРОГАЯ ФИЛЬТРАЦИЯ DASI ТОЛЬКО ПО ПАРТНЕРАМ MARVEL И STN (для стока и SO)
        dasi_df = dasi_df[dasi_df['BP Name'].isin(['Marvel', 'STN'])]
        print(f"Dasi: Оставлены ТОЛЬКО партнеры Marvel и STN. Итого строк: {len(dasi_df)}")
        
        dasi_cols = ['BP Name', 'Part number', 'QTY TOTINV']
        if 'CURR WEEK - 1 SO' in dasi_df.columns:
            dasi_cols.append('CURR WEEK - 1 SO')
        dasi_agg = dasi_df[dasi_cols].groupby(['BP Name', 'Part number']).sum().reset_index()
    else:
        dasi_agg = pd.DataFrame(columns=['BP Name', 'Part number'])
        dasi_specs = pd.DataFrame(columns=['Part number'])
        transit_agg = pd.DataFrame(columns=['BP Name', 'Part number', 'Volume'])

    print("\nЗагрузка Wesi...")
    wesi_path = find_file('b2593d16') or find_file('Wesi')
    if wesi_path:
        wesi_df = pd.read_excel(wesi_path)
        # Wesi: расширенная фильтрация (Consumer + Commercial)
        if 'PH1 Desc' in wesi_df.columns:
            mask = wesi_df['PH1 Desc'].astype(str).str.lower().str.strip().isin(WESI_SEGMENTS)
            wesi_df = wesi_df[mask].copy()
            print(f"Wesi: Отфильтровано по сегментам (PH1 Desc). Осталось строк: {len(wesi_df)}")
        
        # Извлечение спецификаций Wesi (приоритет)
        wesi_spec_cols = {'Product description': 'Description', 'CPU': 'CPU', 'CPU Type': 'CPU type', 'RAM': 'RAM', 'HDD': 'Storage', 'VGA': 'Video', 'OS': 'OS', 'Common_Series': 'Platform', 'PH1 Desc': 'Category'}
        avail_wesi_specs = ['Part number'] + [col for col in wesi_spec_cols.keys() if col in wesi_df.columns]
        wesi_specs = wesi_df[avail_wesi_specs].drop_duplicates('Part number').rename(columns=wesi_spec_cols)
        
        wesi_df['BP Name'] = wesi_df['BP Name'].apply(normalize_partner)
        wesi_cols = ['BP Name', 'Part number', 'Qty TotInv']
        if 'Last Week -1 SO' in wesi_df.columns:
            wesi_cols.append('Last Week -1 SO')
        wesi_agg = wesi_df[wesi_cols].groupby(['BP Name', 'Part number']).sum().reset_index()
        
        # Проверка на неопознанных партнеров
        unknown_sales = wesi_df[(wesi_df['BP Name'] == 'Unknown') | (wesi_df['BP Name'] == 'UNKNOWN')]
        if 'Last Week -1 SO' in unknown_sales.columns:
            unknown_sales = unknown_sales[unknown_sales['Last Week -1 SO'] > 0]
            if not unknown_sales.empty:
                print("\n" + "!"*50)
                print("ВНИМАНИЕ! Найдены продажи по неопознанным партнерам!")
                print("Эти данные НЕ ПОПАДУТ в итоговый отчет, так как имя партнера не найдено в словаре синонимов (partners_map.json).")
                print("Оригинальные имена из файла Wesi:")
                for orig_name in unknown_sales['CUSTNAME'] if 'CUSTNAME' in unknown_sales.columns else unknown_sales.get('Original_BP_Name', unknown_sales.iloc[:, 0].head(5)):
                    # В Wesi колонка изначально называется 'BP Name', но мы ее перезаписали.
                    pass # Чтобы не усложнять, просто предупреждаем пользователя.
                print("Пожалуйста, проверьте исходный файл Wesi!")
                print("Пожалуйста, проверьте исходный файл Wesi!")
                print("!"*50 + "\n")
    else:
        wesi_agg = pd.DataFrame(columns=['BP Name', 'Part number'])
        wesi_specs = pd.DataFrame(columns=['Part number'])

    # Объединяем спецификации (Wesi главнее)
    all_specs = pd.concat([wesi_specs, dasi_specs]).drop_duplicates(subset=['Part number'], keep='first')

    print("\nТранзит загружен из Dasi.")

    print("\nЗагрузка Promo...")
    promo_path = find_file('Promo')
    if promo_path:
        skip = find_header_row(promo_path, 'Part Number')
        promo_df = pd.read_excel(promo_path, skiprows=skip)
        promo_df['Part number'] = promo_df['Part Number']
        gap_col = 'GAP' if 'GAP' in promo_df.columns else ('Support Per unit' if 'Support Per unit' in promo_df.columns else None)
        if gap_col:
            promo_agg = promo_df[['Part number', gap_col]].drop_duplicates('Part number')
        else:
            promo_agg = pd.DataFrame(columns=['Part number'])
    else:
        promo_agg = pd.DataFrame(columns=['Part number'])

    # Создание мастер-списка
    master_keys = pd.concat([
        ds_agg[['BP Name', 'Part number']],
        dasi_agg[['BP Name', 'Part number']],
        wesi_agg[['BP Name', 'Part number']],
        transit_agg[['BP Name', 'Part number']]
    ]).drop_duplicates().dropna(subset=['Part number'])
    
    print(f"\nНайдено уникальных комбинаций Партнер + Партномер за текущую неделю: {len(master_keys)}")

    curr_week_df = master_keys.copy()
    
    # Подтягиваем агрегированные данные
    curr_week_df = curr_week_df.merge(dasi_agg, on=['BP Name', 'Part number'], how='left')
    curr_week_df = curr_week_df.merge(wesi_agg, on=['BP Name', 'Part number'], how='left')
    curr_week_df = curr_week_df.merge(transit_agg, on=['BP Name', 'Part number'], how='left')
    curr_week_df = curr_week_df.merge(ds_agg, on=['BP Name', 'Part number'], how='left')
    curr_week_df = curr_week_df.merge(promo_agg, on='Part number', how='left')
    
    # ПОДТЯГИВАЕМ СПЕЦИФИКАЦИИ
    curr_week_df = curr_week_df.merge(all_specs, on='Part number', how='left')
    
    # Дозаполняем спецификации из исторических данных
    for col in avail_spec_cols:
        if col not in curr_week_df.columns:
            curr_week_df[col] = None
        mask = curr_week_df[col].isna() | (curr_week_df[col].astype(str).str.strip() == '') | (curr_week_df[col].astype(str).str.lower() == 'nan')
        
        def get_hist_val(pn, c):
            val = historical_specs.get(pn, {}).get(c)
            return str(val) if pd.notna(val) else None
            
        curr_week_df.loc[mask, col] = curr_week_df.loc[mask, 'Part number'].map(lambda pn: get_hist_val(pn, col))
    
    # Расчет финальных колонок
    curr_week_df['QTY TOTINV'] = curr_week_df.get('QTY TOTINV', 0).fillna(0)
    curr_week_df['Qty TotInv'] = curr_week_df.get('Qty TotInv', 0).fillna(0)
    curr_week_df['Stock'] = curr_week_df['QTY TOTINV'] + curr_week_df['Qty TotInv']
    
    curr_week_df['Dasi_SO'] = curr_week_df.get('CURR WEEK - 1 SO', 0).fillna(0)
    curr_week_df['Wesi_SO'] = curr_week_df.get('Last Week -1 SO', 0).fillna(0)
    
    # Sales Throught In = продажи из Dasi
    curr_week_df['Sales Throught In'] = curr_week_df['Dasi_SO']
    
    # SO = только Wesi
    curr_week_df['SO'] = curr_week_df['Wesi_SO']
    
    curr_week_df['Transit'] = curr_week_df.get('Volume', 0).fillna(0)
    
    if gap_col and gap_col in curr_week_df.columns:
        curr_week_df['Promo'] = curr_week_df[gap_col]
    else:
        curr_week_df['Promo'] = None

    # Order Load: начинаем с полного QTY из Daily Ship
    curr_week_df['Order Load'] = curr_week_df.get('QTY', 0).fillna(0)
    curr_week_df['Status'] = curr_week_df.get('STATUS', None)
    curr_week_df['SHIP_QTY'] = curr_week_df.get('SHIP_QTY', 0).fillna(0)
    
    # Правило: если SHIP_QTY совпадает с Transit (тот же партнёр, партномер, количество),
    # то вычитаем SHIP_QTY из OL и меняем статус на не-SHIP
    ship_confirmed_by_transit = (
        (curr_week_df['SHIP_QTY'] > 0) & 
        (curr_week_df['Transit'] > 0) & 
        (curr_week_df['SHIP_QTY'] == curr_week_df['Transit'])
    )
    curr_week_df.loc[ship_confirmed_by_transit, 'Order Load'] = curr_week_df.loc[ship_confirmed_by_transit, 'Order Load'] - curr_week_df.loc[ship_confirmed_by_transit, 'SHIP_QTY']
    # Если есть не-SHIP статус, используем его; иначе очищаем
    if 'NON_SHIP_STATUS' in curr_week_df.columns:
        curr_week_df.loc[ship_confirmed_by_transit, 'Status'] = curr_week_df.loc[ship_confirmed_by_transit, 'NON_SHIP_STATUS']
    
    for col in ['Stock', 'SO', 'Sales Throught In', 'Transit', 'Order Load']:
        curr_week_df.loc[curr_week_df[col] == 0, col] = None
        
    curr_week_df['Week'] = week_num
    
    # Добавление столбца Price (с учетом Партнера и Партномера)
    curr_week_df['Price'] = curr_week_df.apply(lambda row: historical_prices.get((row['BP Name'], row['Part number']), None), axis=1)
    
    # Заполнение Segment из исторических данных
    curr_week_df['Segment'] = curr_week_df.apply(lambda row: historical_segments.get((row['BP Name'], row['Part number']), None), axis=1)

    # Заполнение Platform из исторических данных
    curr_week_df['Platform'] = curr_week_df.apply(lambda row: historical_platforms.get((row['BP Name'], row['Part number']), row.get('Platform')), axis=1)

    # Применение логики SubSegment
    curr_week_df['SubSegment'] = curr_week_df.apply(lambda row: get_subsegment(row.get('Category'), row.get('Description'), row.get('CPU type'), row.get('Video')), axis=1)

    # Добавление Processor brand
    def get_processor_brand(cpu_str):
        if pd.isna(cpu_str):
            return None
        c = str(cpu_str).upper()
        if 'AMD' in c or 'RYZEN' in c or 'ATHLON' in c:
            return 'AMD'
        if 'SNAPDRAGON' in c or 'QUALCOMM' in c or 'SD_X' in c:
            return 'Qualcomm'
        if 'INTEL' in c or 'CORE' in c or 'PENTIUM' in c or 'CELERON' in c or 'ULTRA' in c or 'I3' in c or 'I5' in c or 'I7' in c or 'I9' in c:
            return 'Intel'
        # Fallback to Intel as default for undefined
        return 'Intel'
        
    curr_week_df['Processor brand'] = curr_week_df['CPU type'].apply(get_processor_brand)

    final_cols = ['Week', 'BP Name', 'Part number', 'Platform', 'Category', 'Segment', 'SubSegment', 'Description', 'CPU', 'CPU type', 'Processor brand', 'RAM', 'Storage', 'Video', 'OS', 'Price', 'Stock', 'SO', 'Sales Throught In', 'Transit', 'Order Load', 'Status', 'Promo']
    
    # Заполняем пропущенные колонки спецификаций пустотой, если их не было в исходниках
    for col in final_cols:
        if col not in curr_week_df.columns:
            curr_week_df[col] = None
            
    # Форматирование колонки RAM (оставляем только цифры)
    if 'RAM' in curr_week_df.columns:
        curr_week_df['RAM'] = curr_week_df['RAM'].astype(str).str.extract(r'(\d+)')[0]
        curr_week_df['RAM'] = pd.to_numeric(curr_week_df['RAM'], errors='coerce')
        
    # Форматирование колонки OS (Free DOS -> NOOS)
    if 'OS' in curr_week_df.columns:
        def normalize_os(os_val):
            if pd.isna(os_val) or str(os_val).strip().lower() in ['nan', 'none']:
                return None
            val = str(os_val).strip().lower()
            if 'free' in val and 'dos' in val:
                return 'NOOS'
            if 'noos' in val or 'no operating' in val or 'no_operating' in val or ('no' in val and 'os' in val):
                return 'NOOS'
            if 'win' in val:
                return 'Win'
            return str(os_val).strip()
            
        curr_week_df['OS'] = curr_week_df['OS'].apply(normalize_os)
    
    # Форматирование колонки Video (Discrete NVIDIA RTX 3050 -> RTX 3050)
    if 'Video' in curr_week_df.columns:
        import re
        def normalize_video(video_val):
            if pd.isna(video_val) or str(video_val).strip().lower() in ['nan', 'none', '-', '']:
                return 'Int'
            val = str(video_val).strip()
            # Замена всей ячейки на Int, если есть слово Integrated, Vega 8 или HD Graphics
            if re.search(r'(?i)(integrated|vega 8|hd graphics)', val):
                return 'Int'
            
            # Ищем модель GPU (RTX/GTX/MX + номер) в строке с NVIDIA
            if 'nvidia' in val.lower():
                match = re.search(r'((?:RTX|GTX|MX|GT)\s*\d+\w*(?:\s*(?:Ti|SUPER|LHR))?)', val, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
                # Если не нашли, ищем просто числовую модель (например, NVIDIA 4070 SUPER)
                match = re.search(r'(?:nvidia\s+)?(\d{3,4}\w*(?:\s*(?:Ti|SUPER|LHR))?)', val, re.IGNORECASE)
                if match:
                    model = match.group(1).strip()
                    # Если начинается с 3, 4, 5 и является 4-значным числом (RTX 30xx, 40xx, 50xx)
                    if model.startswith(('3', '4', '5')) and len(model.split()[0]) == 4:
                        return f"RTX {model}"
                    return model
            return val
            
        curr_week_df['Video'] = curr_week_df['Video'].apply(normalize_video)
            
    new_data = curr_week_df[final_cols]
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out_path = os.path.join(OUTPUT_DIR, f"WorkPlace_Updated_W{week_num}_{timestamp}.xlsx")
    
    print(f"\nКопирование шаблона WorkPlace и добавление новых данных в {out_path}...")
    try:
        shutil.copy2(wp_path, out_path)
        with pd.ExcelWriter(out_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Сохраняем порядок колонок из WorkPlace
            ordered_cols = list(wp_df.columns)
            
            # Если Category нет, добавляем после Segment/Platform
            if 'Category' not in ordered_cols:
                if 'Segment' in ordered_cols:
                    idx = ordered_cols.index('Segment') + 1
                elif 'Platform' in ordered_cols:
                    idx = ordered_cols.index('Platform') + 1
                else:
                    idx = len(ordered_cols)
                ordered_cols.insert(idx, 'Category')
                
            # Если SubSegment нет, добавляем после Category или Segment
            if 'SubSegment' not in ordered_cols:
                if 'Category' in ordered_cols:
                    idx = ordered_cols.index('Category') + 1
                elif 'Segment' in ordered_cols:
                    idx = ordered_cols.index('Segment') + 1
                else:
                    idx = len(ordered_cols)
                ordered_cols.insert(idx, 'SubSegment')
                
            # Если Processor brand нет, добавляем после CPU type
            if 'Processor brand' not in ordered_cols:
                if 'CPU type' in ordered_cols:
                    idx = ordered_cols.index('CPU type') + 1
                elif 'CPU' in ordered_cols:
                    idx = ordered_cols.index('CPU') + 1
                else:
                    idx = len(ordered_cols)
                ordered_cols.insert(idx, 'Processor brand')

            # Удаляем старые данные за эту же неделю (чтобы избежать дублей)
            if 'Week' in wp_df.columns:
                wp_df_clean = wp_df[wp_df['Week'].astype(str) != str(week_num)].copy()
            else:
                wp_df_clean = wp_df.copy()
            
            # Сохраняем старые данные как есть (Segment уже содержит ручные данные)
            if 'Category' not in wp_df_clean.columns:
                wp_df_clean['Category'] = None
            if 'SubSegment' not in wp_df_clean.columns:
                wp_df_clean['SubSegment'] = None
            if 'Processor brand' not in wp_df_clean.columns:
                wp_df_clean['Processor brand'] = None
            
            # Подгоняем новые данные под колонки из WorkPlace
            for col in ordered_cols:
                if col not in new_data.columns:
                    new_data[col] = None
            
            new_data_ordered = new_data[ordered_cols]
            
            # Объединяем: старые данные (без текущей недели) + новые данные текущей недели
            combined = pd.concat([wp_df_clean[ordered_cols], new_data_ordered], ignore_index=True)
            
            # Жесткая очистка от пустых строк и пробелов перед сохранением
            import numpy as np
            combined.replace(r'^\s*$', np.nan, regex=True, inplace=True)
            combined.dropna(how='all', inplace=True)
            if 'Part number' in combined.columns:
                combined.dropna(subset=['Part number'], inplace=True)
            
            # Очищаем лист и записываем всё заново
            sheet = writer.book[wp_sheet_name]
            sheet.delete_rows(1, sheet.max_row)
            
            combined.to_excel(writer, sheet_name=wp_sheet_name, index=False, header=True, startrow=0)
            
            # Форматирование листа как Таблицы Excel
            try:
                # Очищаем старые таблицы и автофильтры, чтобы не конфликтовать с новой таблицей
                if hasattr(sheet, 'tables') and hasattr(sheet.tables, 'clear'):
                    sheet.tables.clear()
                elif hasattr(sheet, '_tables') and hasattr(sheet._tables, 'clear'):
                    sheet._tables.clear()
                    
                sheet.auto_filter.ref = None
                
                total_rows = len(combined) + 1  # +1 для заголовка
                max_col = len(ordered_cols)
                ref = f"A1:{get_column_letter(max_col)}{total_rows}"
                
                tab = Table(displayName="DataTable", ref=ref)
                
                # Явное определение колонок предотвращает ошибку 'восстановления' в Excel
                for i, col_name in enumerate(ordered_cols, 1):
                    tab.tableColumns.append(TableColumn(id=i, name=str(col_name)))
                
                style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                sheet.add_table(tab)
            except Exception as table_err:
                print(f"Внимание: Не удалось применить стиль таблицы: {table_err}")
            
        print("ГОТОВО! Отчет успешно сформирован, новые данные добавлены к истории, данные отформатированы как Таблица Excel.")
    except Exception as e:
        print(f"Ошибка при сохранении: {e}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Ошибка: Необходимо указать номер недели!")
        sys.exit(1)
    main(sys.argv[1])
