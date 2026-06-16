# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import openpyxl
import re

# Пути к файлам
template_path = r'C:\Users\DenisG\Documents\Workplace\Data\Prices\Line up monitoring.xlsx'
matrix_path = r'C:\Users\DenisG\Documents\Workplace\Data\Prices\Матрица цен конкурентов - все Т2.xlsx'
wp_path = r'C:\Users\DenisG\Documents\Workplace\Data\WorkPlace.xlsx'
output_path = r'C:\Users\DenisG\Documents\Workplace\Data\Prices\Line up monitoring_filled.xlsx'

# LFR столбцы (крупный ритейл)
lfr_columns = [
    'Алсер', 'Белый ветер (Караганда)', 'Мечта (Астана)',
    'Сулпак (Алматы)', 'Технодом (Алматы)', 'Эврика (Шымкент)',
    'DNS (Нур-Султан)'
]

def clean_model_name(name):
    """Вырезаем кавычки из названия модели"""
    if pd.isna(name):
        return 'N/A'
    return str(name).replace('"', '').replace("'", "").strip()

def normalize_text(text):
    if pd.isna(text) or str(text).strip() == "":
        return ""
    text = str(text).lower()
    text = text.replace('core ', '') 
    text = text.replace('gb', '').replace('g', '').replace('tb', 't')
    text = re.sub(r'\s+', '', text)
    return text

def is_match(template_val, matrix_val):
    if pd.isna(template_val) or str(template_val).strip() == "":
        return True
    if pd.isna(matrix_val):
        return False
    t_str = normalize_text(template_val)
    m_str = normalize_text(matrix_val)
    return t_str in m_str

try:
    df_matrix = pd.read_excel(matrix_path)
    # Фильтрация категории
    df_matrix = df_matrix[df_matrix.iloc[:, 0] == 'Ноутбуки/Ультрабуки']
    
    df_wp = pd.read_excel(wp_path, sheet_name='Data')
    df_wp['Stock'] = pd.to_numeric(df_wp['Stock'], errors='coerce').fillna(0)
    stock_map = df_wp.groupby('Part number')['Stock'].sum().to_dict()
    
    start_idx = list(df_matrix.columns).index('RRP') + 1 if 'RRP' in df_matrix.columns else 14
    all_price_cols = df_matrix.columns[start_idx:]
    available_lfr_cols = [c for c in lfr_columns if c in df_matrix.columns]
    
    col_model = df_matrix.columns[1]
    col_brand = df_matrix.columns[2]
    col_disp = df_matrix.columns[3]
    col_cpu = df_matrix.columns[4]
    col_ram = df_matrix.columns[5]
    col_hdd = df_matrix.columns[6]
    col_video = df_matrix.columns[7]
    col_os = df_matrix.columns[9]
    col_artikul = 'Артикул'
    col_mpn = 'MPN'
    
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    current_brand = None
    
    for row_idx in range(2, sheet.max_row + 1):
        try:
            brand_cell = sheet.cell(row=row_idx, column=1).value
            if brand_cell and str(brand_cell).strip() != "":
                current_brand = str(brand_cell).strip()
                
            price_type = sheet.cell(row=row_idx, column=2).value
            if not price_type or str(price_type).strip() not in ['Lowest Price', 'Lowest Price LFR']:
                continue
                
            display = sheet.cell(row=row_idx, column=5).value
            cpu = sheet.cell(row=row_idx, column=6).value
            ram = sheet.cell(row=row_idx, column=7).value
            storage = sheet.cell(row=row_idx, column=8).value
            video = sheet.cell(row=row_idx, column=9).value
            os_val = sheet.cell(row=row_idx, column=10).value
            
            # Фильтрация
            mask_brand = df_matrix[col_brand].apply(lambda x: is_match(current_brand, x))
            mask_disp = df_matrix[col_disp].apply(lambda x: is_match(display, x))
            mask_cpu = df_matrix[col_cpu].apply(lambda x: is_match(cpu, x))
            mask_ram = df_matrix[col_ram].apply(lambda x: is_match(ram, x))
            mask_hdd = df_matrix[col_hdd].apply(lambda x: is_match(storage, x))
            mask_video = df_matrix[col_video].apply(lambda x: is_match(video, x))
            mask_os = df_matrix[col_os].apply(lambda x: is_match(os_val, x))
            
            matched = df_matrix[mask_brand & mask_disp & mask_cpu & mask_ram & mask_hdd & mask_video & mask_os].copy()
            
            if matched.empty:
                sheet.cell(row=row_idx, column=3).value = 'N/A'
                sheet.cell(row=row_idx, column=4).value = 'N/A'
                continue
                
            if current_brand and current_brand.lower() == 'lenovo':
                artikul_col = col_artikul if col_artikul in matched.columns else matched.columns[10]
                matched['stock_val'] = matched[artikul_col].apply(lambda x: stock_map.get(x, 0))
                max_stock = matched['stock_val'].max()
                matched = matched[matched['stock_val'] == max_stock]
                
            best_price = np.inf
            best_model = 'N/A'
            
            target_cols = all_price_cols if str(price_type).strip() == 'Lowest Price' else available_lfr_cols
            
            for _, m_row in matched.iterrows():
                prices = pd.to_numeric(m_row[target_cols], errors='coerce')
                prices = prices.replace(0, np.nan).replace(0.0, np.nan)
                min_p = prices.min(skipna=True)
                
                if pd.notna(min_p) and min_p < best_price:
                    best_price = float(min_p)
                    model_val = m_row.get(col_model, m_row.get(col_mpn, 'N/A'))
                    best_model = clean_model_name(model_val)
                    
            if best_price == np.inf:
                sheet.cell(row=row_idx, column=3).value = 'N/A'
                sheet.cell(row=row_idx, column=4).value = 'N/A'
            else:
                sheet.cell(row=row_idx, column=3).value = best_price
                sheet.cell(row=row_idx, column=4).value = best_model
                
        except Exception as e:
            sheet.cell(row=row_idx, column=3).value = 'N/A'
            sheet.cell(row=row_idx, column=4).value = 'N/A'
            
    wb.save(output_path)
    print("Done")
except Exception as e:
    print("ERROR:", e)
