import openpyxl
import sys
import os

file_path = r'C:\Users\DenisG\Documents\Workplace\Data\GFK\GFK data TTL.xlsx'

ALLOWED_SEGMENTS = {
    'Apple', 'Basic Celeron', 'Basic i3R3', 'Basic i5R5', 'Basic i7R7', 'Basic i9R9', 
    'Basic Pentium Athlon', 'Gaming i5R5 50series', 'Gaming i5R5 60series', 'Gaming i7R7 50series', 
    'Gaming i7R7 60series', 'Gaming i7R7 70series', 'Gaming i9R9 60series', 'Gaming i9R9 70series', 
    'Gaming i9R9 80+series', 'Gaming Others', 'Premium i3R3', 'Premium i5R5', 'Premium i7R7', 
    'Premium i9R9', 'Snapdragon'
}

ALLOWED_SUBSEGMENTS = {
    'Apple', 'Basic Celeron', 'Basic Pentium Athlon', 'Basic R3', 'Basic i3', 'Basic R5', 'Basic i5', 
    'Basic R7', 'Basic i7', 'Basic R9', 'Basic i9', 'Premium i3R3', 'Premium i5R5', 'Premium i7R7', 
    'Premium i9R9', 'SnapDr Laptops', 'Gaming i5R5 3050', 'Gaming i7R7 3050', 'Gaming i5R5 3060', 
    'Gaming i7R7 3060', 'Gaming i7R7 3070+', 'Gaming i5R5 4050', 'Gaming i7R7 4050', 'Gaming i5R5 4060', 
    'Gaming i7R7 4060', 'Gaming i7R7 4070+', 'Gaming i9R9 4070+', 'Gaming i9R9 4080+', 'Gaming i5R5 5050', 
    'Gaming i7R7 5050', 'Gaming i5R5 5060', 'Gaming i7R7 5060', 'Gaming i9R9 5060', 'Gaming i7R7 5070+', 
    'Gaming i9R9 5070+', 'Gaming Others'
}

print("Loading workbook with openpyxl (this may take a minute)...")
try:
    wb = openpyxl.load_workbook(file_path)
except Exception as e:
    print("Failed to load workbook:", e)
    sys.exit(1)

if 'Data' not in wb.sheetnames:
    print("Sheet 'Data' not found!")
    sys.exit(1)

ws = wb['Data']

# Find column indices (1-based)
headers = {}
for cell in ws[1]:
    if cell.value:
        headers[str(cell.value).strip()] = cell.column

required = ['Brand', 'Processor', 'Gaming PCs', 'Model Name', 'GPU Model']
for req in required:
    if req not in headers:
        print(f"Required column '{req}' not found on 'Data' sheet!")
        sys.exit(1)

# Ensure Tech_Segment and Tech_SubSegment exist
if 'Tech_Segment' not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value='Tech_Segment')
    headers['Tech_Segment'] = ws.max_column
if 'Tech_SubSegment' not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value='Tech_SubSegment')
    headers['Tech_SubSegment'] = ws.max_column

idx_brand = headers['Brand']
idx_proc = headers['Processor']
idx_gaming = headers['Gaming PCs']
idx_model = headers['Model Name']
idx_gpu = headers['GPU Model']
idx_seg = headers['Tech_Segment']
idx_sub = headers['Tech_SubSegment']

def get_mapping(brand, processor, gaming_pc, model_name, gpu_model):
    if brand == 'apple': return 'Apple', 'Apple'
    if 'snapdragon' in processor or 'snapd.' in processor: return 'Snapdragon', 'SnapDr Laptops'

    cpu_segment = ''
    cpu_subsegment = ''
    if 'core i3' in processor or 'core 3' in processor or 'core ultra 3' in processor or 'u-series' in processor: cpu_segment, cpu_subsegment = 'i3R3', 'i3'
    elif 'ryzen 3' in processor or 'ryzen ai 3' in processor or 'ryzen ai 300' in processor: cpu_segment, cpu_subsegment = 'i3R3', 'R3'
    elif 'core i5' in processor or 'core 5' in processor or 'core ultra 5' in processor: cpu_segment, cpu_subsegment = 'i5R5', 'i5'
    elif 'ryzen 5' in processor or 'ryzen ai 5' in processor: cpu_segment, cpu_subsegment = 'i5R5', 'R5'
    elif 'core i7' in processor or 'core 7' in processor or 'core ultra 7' in processor: cpu_segment, cpu_subsegment = 'i7R7', 'i7'
    elif 'ryzen 7' in processor or 'ryzen ai 7' in processor: cpu_segment, cpu_subsegment = 'i7R7', 'R7'
    elif 'core i9' in processor or 'core 9' in processor or 'core ultra 9' in processor: cpu_segment, cpu_subsegment = 'i9R9', 'i9'
    elif 'ryzen 9' in processor or 'ryzen ai 9' in processor: cpu_segment, cpu_subsegment = 'i9R9', 'R9'
    elif 'celeron' in processor or 'n-series' in processor or 'atom' in processor: cpu_segment, cpu_subsegment = 'Celeron', 'Celeron'
    elif 'pentium' in processor or 'athlon' in processor or '3000 series' in processor or 'a4-series' in processor or 'a6-series' in processor or 'a9-series' in processor: cpu_segment, cpu_subsegment = 'Pentium Athlon', 'Pentium Athlon'

    gaming_models = ['legion', 'rog', 'tuf', 'omen', 'nitro', 'alienware', 'predator', 'loq']
    is_gaming = (gaming_pc == 'gaming') or any(m in model_name for m in gaming_models)
    
    if is_gaming:
        if not cpu_segment: 
            return 'Gaming Others', 'Gaming Others'

        gpu_seg, gpu_sub = '', ''
        if '3050' in gpu_model: gpu_seg, gpu_sub = '50series', '3050'
        elif '3060' in gpu_model: gpu_seg, gpu_sub = '60series', '3060'
        elif any(x in gpu_model for x in ['3070', '3080', '3090']): gpu_seg, gpu_sub = '70series', '3070+'
        elif '4050' in gpu_model: gpu_seg, gpu_sub = '50series', '4050'
        elif '4060' in gpu_model: gpu_seg, gpu_sub = '60series', '4060'
        elif '4070' in gpu_model: gpu_seg, gpu_sub = '70series', '4070+'
        elif any(x in gpu_model for x in ['4080', '4090']): gpu_seg, gpu_sub = '80+series', '4080+'
        elif '5050' in gpu_model: gpu_seg, gpu_sub = '50series', '5050'
        elif '5060' in gpu_model: gpu_seg, gpu_sub = '60series', '5060'
        elif any(x in gpu_model for x in ['5070', '5080', '5090']): gpu_seg, gpu_sub = '70series', '5070+'
        
        if not gpu_seg: return 'Gaming Others', 'Gaming Others'
        return f'Gaming {cpu_segment} {gpu_seg}', f'Gaming {cpu_segment} {gpu_sub}'

    premium_models = ['zenbook', 'spectre', 'xps', 'yoga', 'envy', 'elitebook', 'swift']
    if any(m in model_name for m in premium_models):
        if not cpu_segment: return '', ''
        return f'Premium {cpu_segment}', f'Premium {cpu_segment}'

    if not cpu_segment: return '', ''
    return f'Basic {cpu_segment}', f'Basic {cpu_subsegment}'

print("Processing rows...")
blanks_left = 0
cells_updated = 0

for row in range(2, ws.max_row + 1):
    brand = str(ws.cell(row=row, column=idx_brand).value or '').lower().strip()
    processor = str(ws.cell(row=row, column=idx_proc).value or '').lower().strip()
    gaming_pc = str(ws.cell(row=row, column=idx_gaming).value or '').lower().strip()
    model_name = str(ws.cell(row=row, column=idx_model).value or '').lower().strip()
    gpu_model = str(ws.cell(row=row, column=idx_gpu).value or '').lower().strip()

    seg, sub = get_mapping(brand, processor, gaming_pc, model_name, gpu_model)

    if seg not in ALLOWED_SEGMENTS: 
        seg = 'Gaming Others' if seg.startswith('Gaming ') else ""
    if sub not in ALLOWED_SUBSEGMENTS: 
        sub = 'Gaming Others' if sub.startswith('Gaming ') else ""

    # ВАЖНО: Если скрипт знает ответ, мы его пишем. 
    # Если скрипт НЕ знает ответ (пустота), мы НЕ перезаписываем ячейку, 
    # чтобы сохранить ручные правки пользователя!
    
    current_seg = ws.cell(row=row, column=idx_seg).value
    current_sub = ws.cell(row=row, column=idx_sub).value
    
    if seg != "" and current_seg != seg:
        ws.cell(row=row, column=idx_seg, value=seg)
        cells_updated += 1
    elif seg == "" and not current_seg:
        blanks_left += 1
        
    if sub != "" and current_sub != sub:
        ws.cell(row=row, column=idx_sub, value=sub)

print("Saving workbook...")
try:
    wb.save(file_path)
    print("Successfully saved:", file_path)
    print(f"Cells updated with new mappings: {cells_updated}")
    print(f"Rows still blank: {blanks_left}")
except Exception as e:
    print("Error saving:", e)
    sys.exit(1)
